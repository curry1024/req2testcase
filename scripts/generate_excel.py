"""
测试用例 Excel 生成脚本
输入: step4_testcases.json
输出: step4_testcases.xlsx
表头: 用例编号, 模块, 测试点, 用例标题, 前置条件, 测试步骤, 预期结果, 优先级, 测试结果, 备注

用法: python generate_excel.py <testcases_json_path> [output_path]
"""
import json
import sys
import os

HEADERS = [
    "用例编号", "模块", "测试点", "用例标题",
    "前置条件", "测试步骤", "预期结果", "优先级",
    "测试结果", "备注"
]

COL_WIDTHS = {
    "用例编号": 12, "模块": 12, "测试点": 22, "用例标题": 28,
    "前置条件": 28, "测试步骤": 40,
    "预期结果": 30, "优先级": 8, "测试结果": 10, "备注": 20
}

CENTER_COLS = {"用例编号", "模块", "测试点", "优先级", "测试结果"}


def merge_module_cells(ws, start_row, end_row, col_idx):
    """合并相同模块的单元格"""
    from openpyxl.styles import Alignment
    if end_row <= start_row:
        return
    merge_start = start_row
    prev_module = ws.cell(row=start_row, column=col_idx).value
    for r in range(start_row + 1, end_row + 1):
        current = ws.cell(row=r, column=col_idx).value
        if current != prev_module:
            if r - 1 > merge_start:
                ws.merge_cells(start_row=merge_start, start_column=col_idx,
                               end_row=r - 1, end_column=col_idx)
            merge_start = r
            prev_module = current
    if end_row > merge_start:
        ws.merge_cells(start_row=merge_start, start_column=col_idx,
                       end_row=end_row, end_column=col_idx)


def generate_excel(json_path, output_path=None):
    if not os.path.exists(json_path):
        print(f"[ERROR] File not found: {json_path}", file=sys.stderr)
        sys.exit(1)

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    testcases = data.get("testcases", [])

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 样式定义
    header_font = Font(name="微软雅黑", size=11, bold=True, color="000000")
    header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_font = Font(name="微软雅黑", size=10)
    body_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="B4B4B4"),
        right=Side(style="thin", color="B4B4B4"),
        top=Side(style="thin", color="B4B4B4"),
        bottom=Side(style="thin", color="B4B4B4"),
    )

    # 数据排序：按模块 + 编号
    testcases.sort(key=lambda tc: (tc.get("module", ""), tc.get("id", "")))

    # 写入表头
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    for row_idx, tc in enumerate(testcases, 2):
        steps = tc.get("steps", [])
        steps_text = "\n".join(f"{i+1}. {s}" for i, s in enumerate(steps)) if steps else ""

        preconditions = tc.get("preconditions", [])
        pre_text = "\n".join(f"- {p}" for p in preconditions) if preconditions else ""

        values = [
            tc.get("id", ""),
            tc.get("module", ""),
            tc.get("feature_point", ""),
            tc.get("title", ""),
            pre_text,
            steps_text,
            tc.get("expected_result", ""),
            tc.get("priority", ""),
            tc.get("test_result", ""),
            tc.get("notes", ""),
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.border = thin_border

            cell.alignment = body_alignment

    # 合并相同模块 + 相同测试点单元格（居中）
    if len(testcases) > 0:
        merge_module_cells(ws, 2, 1 + len(testcases), 2)
        merge_module_cells(ws, 2, 1 + len(testcases), 3)

    # 设置列宽
    for col_idx, header in enumerate(HEADERS, 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = COL_WIDTHS.get(header, 15)

    # 冻结首行 + 首列
    ws.freeze_panes = "C2"

    # 添加筛选
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(testcases)+1}"

    if output_path is None:
        output_path = json_path.replace(".json", ".xlsx")

    wb.save(output_path)

    total = data.get("meta", {}).get("total_testcases", len(testcases))
    print(f"[OK] Excel generated: {output_path}")
    print(f"[INFO] Total test cases: {total}")
    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python generate_excel.py <testcases_json_path> [output_path]", file=sys.stderr)
        sys.exit(1)

    out = sys.argv[2] if len(sys.argv) > 2 else None
    generate_excel(sys.argv[1], out)
