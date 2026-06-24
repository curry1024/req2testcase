"""
最终测试用例 Excel 生成脚本
支持多 Sheet、条件着色、下拉选项、汇总统计

用法:
  # 自动按模块分 Sheet
  python generate_final_excel.py --tc <testcases.json> --output <文件名>

  # LLM 指定分 Sheet 方案
  python generate_final_excel.py --tc <testcases.json> --config <sheet_config.json> --output <文件名>
"""
import json
import sys
import os
import re
from datetime import datetime

HEADERS = [
    "用例编号", "模块", "测试点", "用例标题",
    "前置条件", "测试步骤", "预期结果", "优先级",
    "测试结果", "备注"
]

COL_WIDTHS = {
    "用例编号": 12, "模块": 12, "测试点": 22, "用例标题": 28,
    "前置条件": 28, "测试步骤": 40,
    "预期结果": 30, "优先级": 8, "测试结果": 10, "备注": 20
}

CENTER_COLS = {"用例编号", "模块", "测试点", "优先级", "测试结果"}

PRIORITY_FILLS = {
    "P0": "FCB9B2",
    "P1": "FCD6B2",
    "P2": "FCE4B2",
    "P3": "E2E2E2",
}

RESULT_FILL = "C6EFCE"


def load_json(filepath):
    with open(filepath, "r", encoding="utf-8") as f:
        return json.load(f)


def build_sheet_groups(testcases, config_path=None):
    """构建 Sheet 分组。有 config 用 config，无则自动按模块分。"""
    if config_path and os.path.exists(config_path):
        config = load_json(config_path)
        groups = config.get("sheet_groups", [])

        grouped_ids = set()
        for g in groups:
            grouped_ids.update(g.get("testcase_ids", []))

        all_ids = {tc["id"] for tc in testcases}
        ungrouped = all_ids - grouped_ids
        if ungrouped:
            groups.append({
                "sheet_name": "其他",
                "testcase_ids": sorted(ungrouped),
            })

        return groups

    # 自动按模块分组
    module_map = {}
    for tc in testcases:
        mod = tc.get("module", "未分类")
        if mod not in module_map:
            module_map[mod] = []
        module_map[mod].append(tc["id"])

    return [
        {"sheet_name": mod, "testcase_ids": ids}
        for mod, ids in module_map.items()
    ]


def tc_by_id(testcases):
    return {tc["id"]: tc for tc in testcases}


def as_list(value):
    """归一化为列表，防止字符串被逐字符遍历。"""
    if value is None:
        return []
    if isinstance(value, str):
        value = value.strip()
        return [value] if value else []
    if isinstance(value, (list, tuple)):
        return [v for v in value]
    return [value]


def merge_value_runs(ws, start_row, keys, col_idx):
    """按 keys 中的连续相同值合并 col_idx 列。keys[i] 对应行 start_row+i。"""
    n = len(keys)
    if n == 0:
        return
    run_start = 0
    for i in range(1, n + 1):
        if i == n or keys[i] != keys[run_start]:
            if i - 1 > run_start:
                ws.merge_cells(start_row=start_row + run_start, start_column=col_idx,
                               end_row=start_row + i - 1, end_column=col_idx)
            run_start = i


def merge_module_cells(ws, start_row, end_row, col_idx):
    """合并相同模块的单元格"""
    if end_row <= start_row:
        return
    merge_start = start_row
    prev_module = ws.cell(row=start_row, column=col_idx).value
    for r in range(start_row + 1, end_row + 1):
        current = ws.cell(row=r, column=col_idx).value
        if current != prev_module:
            if r - 1 > merge_start:
                ws.merge_cells(start_row=merge_start, start_column=col_idx,
                               end_row=r - 1, end_column=col_idx)
            merge_start = r
            prev_module = current
    if end_row > merge_start:
        ws.merge_cells(start_row=merge_start, start_column=col_idx,
                       end_row=end_row, end_column=col_idx)


def apply_header_style(ws, col_count):
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    header_font = Font(name="微软雅黑", size=11, bold=True, color="000000")
    header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="B4B4B4"),
        right=Side(style="thin", color="B4B4B4"),
        top=Side(style="thin", color="B4B4B4"),
        bottom=Side(style="thin", color="B4B4B4"),
    )

    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def write_testcase_sheet(ws, testcases, tc_ids, tc_map):
    """写入用例数据 Sheet"""
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    body_font = Font(name="微软雅黑", size=10)
    body_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="B4B4B4"),
        right=Side(style="thin", color="B4B4B4"),
        top=Side(style="thin", color="B4B4B4"),
        bottom=Side(style="thin", color="B4B4B4"),
    )

    # 写入表头
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)

    apply_header_style(ws, len(HEADERS))

    # 筛选该 Sheet 的用例，按模块排序
    sheet_tcs = [tc_map[tid] for tid in tc_ids if tid in tc_map]
    sheet_tcs.sort(key=lambda tc: (tc.get("module", ""), tc.get("id", "")))

    # ---- Build flat row list (multi-row TCs expand to multiple rows) ----
    flat_rows = []
    for tc in sheet_tcs:
        expected_items = tc.get("expected_results", [tc.get("expected_result", "")])
        if not isinstance(expected_items, list) or not expected_items:
            expected_items = [tc.get("expected_result", "")]
        steps = as_list(tc.get("steps"))
        steps_text = "\n".join(f"{i+1}. {s}" for i, s in enumerate(steps)) if steps else ""
        preconditions = as_list(tc.get("preconditions"))
        pre_text = "\n".join(f"- {p}" for p in preconditions) if preconditions else ""
        for i, er in enumerate(expected_items):
            flat_rows.append({
                "id": tc["id"], "module": tc.get("module",""), "feature_point": tc.get("feature_point",""),
                "title": tc.get("title",""), "pre_text": pre_text, "steps_text": steps_text,
                "expected_text": er, "priority": tc.get("priority",""),
                "test_result": tc.get("test_result",""), "notes": tc.get("notes",""),
                "row_count": len(expected_items), "sub": i,
            })

    # ---- Write all rows ----
    for i, row in enumerate(flat_rows):
        r = 2 + i
        values = [row["id"], row["module"], row["feature_point"], row["title"],
                  row["pre_text"], row["steps_text"], row["expected_text"],
                  row["priority"], row["test_result"], row["notes"]]
        for ci, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = body_align
        if row["priority"] in PRIORITY_FILLS:
            pf = PatternFill(start_color=PRIORITY_FILLS[row["priority"]], end_color=PRIORITY_FILLS[row["priority"]], fill_type="solid")
            for ci in range(1, len(HEADERS)+1):
                ws.cell(row=r, column=ci).fill = pf
        ws.cell(row=r, column=9).fill = PatternFill(start_color=RESULT_FILL, end_color=RESULT_FILL, fill_type="solid")

    # ---- Merge cells for multi-row TCs (排除模块/测试点列，由下方按值合并统一处理) ----
    row_offset = 1
    for tc in sheet_tcs:
        n = len(tc.get("expected_results", [tc.get("expected_result","")])) if isinstance(tc.get("expected_results",[]),list) else 1
        if n > 1:
            start_r, end_r = row_offset + 1, row_offset + n
            # 仅合并用例级字段(编号/标题/前置/步骤)；优先级/测试结果/备注按预期结果逐行独立
            for col_m in [1,4,5,6]:
                ws.merge_cells(start_row=start_r, start_column=col_m, end_row=end_r, end_column=col_m)
        row_offset += n

    total_rows = 1 + len(flat_rows)

    # Merge module + feature_point cells（按源数据值合并，避免读取已合并单元格返回 None 造成错误断裂）
    module_keys = [row["module"] for row in flat_rows]
    feature_keys = [(row["module"], row["feature_point"]) for row in flat_rows]
    merge_value_runs(ws, 2, module_keys, 2)
    merge_value_runs(ws, 2, feature_keys, 3)

    # Column widths
    for col_idx, header in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(header, 15)

    # Freeze
    ws.freeze_panes = "C2"

    # Auto filter
    if total_rows > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{total_rows}"

    # 测试结果下拉验证
    dv = DataValidation(type="list", formula1='"P,F,N/A"', allow_blank=True)
    dv.error = "请选择 P, F 或 N/A"
    dv.errorTitle = "无效输入"
    if total_rows > 1:
        dv.add(f"I2:I{total_rows}")
        ws.add_data_validation(dv)

    return len(sheet_tcs)


def write_summary_sheet(ws, testcases, groups, tc_map):
    """写入汇总统计 Sheet"""
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.chart import BarChart3D, Reference
    from openpyxl.utils import get_column_letter

    title_font = Font(name="微软雅黑", size=14, bold=True)
    sub_font = Font(name="微软雅黑", size=11, bold=True)
    body_font = Font(name="微软雅黑", size=10)

    total = len(testcases)
    ws.cell(row=1, column=1, value="测试用例汇总").font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    ws.cell(row=3, column=1, value="生成时间").font = sub_font
    ws.cell(row=3, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M")).font = body_font

    ws.cell(row=4, column=1, value="用例总数").font = sub_font
    ws.cell(row=4, column=2, value=total).font = body_font

    ws.cell(row=5, column=1, value="Sheet 数量").font = sub_font
    ws.cell(row=5, column=2, value=len(groups)).font = body_font

    # 按 Sheet 分布
    ws.cell(row=7, column=1, value="Sheet 分布").font = sub_font
    ws.cell(row=8, column=1, value="Sheet 名称").font = sub_font
    ws.cell(row=8, column=2, value="用例数").font = sub_font

    for i, g in enumerate(groups):
        name = g["sheet_name"]
        count = sum(1 for tid in g["testcase_ids"] if tid in tc_map)
        ws.cell(row=9 + i, column=1, value=name).font = body_font
        ws.cell(row=9 + i, column=2, value=count).font = body_font

    # 按优先级分布
    priority_counts = {"P0": 0, "P1": 0, "P2": 0, "P3": 0}
    for tc in testcases:
        p = tc.get("priority", "")
        if p in priority_counts:
            priority_counts[p] += 1

    priority_start = 9 + len(groups) + 2
    ws.cell(row=priority_start, column=1, value="优先级分布").font = sub_font
    ws.cell(row=priority_start + 1, column=1, value="优先级").font = sub_font
    ws.cell(row=priority_start + 1, column=2, value="数量").font = sub_font

    for i, (p, c) in enumerate(priority_counts.items()):
        row = priority_start + 2 + i
        cell = ws.cell(row=row, column=1, value=p)
        cell.font = body_font
        if p in PRIORITY_FILLS:
            cell.fill = PatternFill(start_color=PRIORITY_FILLS[p], end_color=PRIORITY_FILLS[p], fill_type="solid")
        ws.cell(row=row, column=2, value=c).font = body_font

    # 列宽
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16


def generate_final_excel(tc_path, sheet_config_path=None, output_path=None):
    if not os.path.exists(tc_path):
        print(f"[ERROR] Testcases file not found: {tc_path}", file=sys.stderr)
        sys.exit(1)

    testcases_data = load_json(tc_path)
    testcases = testcases_data.get("testcases", [])
    tc_map = tc_by_id(testcases)

    groups = build_sheet_groups(testcases, sheet_config_path)

    from openpyxl import Workbook

    wb = Workbook()

    # 汇总 Sheet（放第一个）
    ws_summary = wb.active
    ws_summary.title = "汇总"
    write_summary_sheet(ws_summary, testcases, groups, tc_map)

    # 用例 Sheet
    for g in groups:
        ws = wb.create_sheet(title=g["sheet_name"][:31])
        write_testcase_sheet(ws, testcases, g["testcase_ids"], tc_map)

    if output_path is None:
        base = os.path.splitext(os.path.basename(tc_path))[0]
        output_path = base.replace("step4_testcases", "测试用例") + ".xlsx"

    # 版本号处理
    output_path = resolve_version(output_path)

    wb.save(output_path)

    total = len(testcases)
    print(f"[OK] Final Excel: {output_path}")
    print(f"[INFO] Total cases: {total}, Sheets: {len(groups) + 1}")
    for g in groups:
        count = sum(1 for tid in g["testcase_ids"] if tid in tc_map)
        print(f"  - {g['sheet_name']}: {count} 条")


def resolve_version(output_path):
    """处理版本号：v1 → v2 自动递增"""
    if not os.path.exists(output_path):
        return output_path

    match = re.match(r"(.+_v)(\d+)(\.xlsx)$", output_path)
    if match:
        base = match.group(1)
        ver = int(match.group(2))
        ext = match.group(3)
        while True:
            ver += 1
            new_path = f"{base}{ver}{ext}"
            if not os.path.exists(new_path):
                return new_path
    else:
        name, ext = os.path.splitext(output_path)
        ver = 1
        while True:
            ver += 1
            new_path = f"{name}_v{ver}{ext}"
            if not os.path.exists(new_path):
                return new_path


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Generate final testcase Excel")
    parser.add_argument("--tc", required=True, help="Path to step4_testcases.json")
    parser.add_argument("--config", default=None, help="Path to sheet config JSON (LLM generated)")
    parser.add_argument("--output", default=None, help="Output Excel path")
    args = parser.parse_args()

    generate_final_excel(args.tc, args.config, args.output)
